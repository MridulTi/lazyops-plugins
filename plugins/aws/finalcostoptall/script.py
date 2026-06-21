import os
import sys
#!/usr/bin/python3
import boto3
from openpyxl import Workbook
from datetime import datetime, timezone, timedelta
from openpyxl.styles import Font
import re
from botocore.exceptions import ClientError
from datetime import datetime, timedelta
import argparse

# Create an argument parser
parser = argparse.ArgumentParser(description='Script description')

# Define a command-line argument for the profile_name
parser.add_argument('--profile_name', type=str, help='AWS profile name')
#parser.add_argument('--npnskey', type=str, help='Non Prod Not Nightly Shutdown Tag Key')
#parser.add_argument('--npnsvalue', type=str, help='Non Prod Not Nightly Shutdown Tag Value')

# Parse the command-line arguments
args = parser.parse_args()

# Access the value of the profile_name from the command line
profile_name = args.profile_name
#npnskey = args.npnskey
#npnsvalue = args.npnsvalue

# Set up the AWS session using the provided profile_name
boto3.setup_default_session(profile_name=profile_name)

region_name = os.environ.get("AWS_REGION") or os.environ.get("AWS_DEFAULT_REGION") or sys.argv[1] if len(sys.argv) > 1 else None



client = boto3.client('ec2', region_name=region_name)

wb = Workbook()
file_name = f'{profile_name}-cost-optimisation.xlsx'

#Non Prod Not Nightly Shutdown
#ec2_client = boto3.client('ec2', region_name=region_name)
#response = ec2_client.describe_instances()
#
# Extract instance IDs
#instance_info = []
#for reservation in response['Reservations']:
#    for instance in reservation['Instances']:
#        env_tag_value = None
#        schedule_tag = None
#        for tag in instance.get("Tags", []):
#            if tag["Key"] == "Stack":
#                env_tag_value = tag["Value"]
#            elif tag["Key"] == "{npnskey}":
#                schedule_tag = tag["Value"]
#        if (env_tag_value is None or env_tag_value != "prod") and (schedule_tag is None or schedule_tag == "{npnsvalue}"):
#            instance_info.append({
#               'Instance Id': instance['InstanceId'],
#               'instance Name': next((tag['Value'] for tag in instance.get('Tags', []) if tag['Key'] == 'Name'), None),
#               'Instance Type': instance['InstanceType'],
#               'Private Ip Address': instance.get('PrivateIpAddress', 'Key not found'),
#               'Techteam': next((tag['Value'] for tag in instance.get('Tags', []) if tag['Key'] == 'Techteam'), None)
#            })
#
#npnns = instance_info

#Stopped Instances
client = boto3.client('ec2', region_name=region_name)
response = client.describe_instances()
stopped_instances=[]
for reservation in response['Reservations']:
    for instance in reservation['Instances']:
        if instance['State']['Name'] == 'stopped':
           stopped_instances.append({
             'Name': next((tag['Value'] for tag in instance.get('Tags', []) if tag['Key'] == 'Name'), None),
             'Instance Id': instance['InstanceId'],
             'Instance Type': instance['InstanceType'],
             'Private Ip Address': instance['PrivateIpAddress'],
             'Techteam': next((tag['Value'] for tag in instance.get('Tags', []) if tag['Key'] == 'Techteam'), None)
           })
stopint = stopped_instances

#Overprovisioned Instances
client = boto3.client('compute-optimizer', region_name=region_name)
ec2_client = boto3.client('ec2', region_name=region_name)

response1 = client.get_ec2_instance_recommendations(
)

over = []

if 'instanceRecommendations' in response1:
   for recommendation in response1['instanceRecommendations']:
       if recommendation['finding'] == 'OVER_PROVISIONED':
          over.append(recommendation['instanceArn'])

savings = []
response = client.get_ec2_instance_recommendations(
    instanceArns=over,
    maxResults=123,
    recommendationPreferences={
        'cpuVendorArchitectures': [
            'AWS_ARM64',
            'CURRENT',
        ]
    }
)

if 'instanceRecommendations' in response:
   for recommendation in response['instanceRecommendations']:
       instance_id = recommendation['instanceArn'].split('/')[1]
       instance_info = ec2_client.describe_instances(InstanceIds=[instance_id])
       if 'Reservations' in instance_info:
          for reservation in instance_info['Reservations']:
              for instance in reservation['Instances']:
                  Prvipaddr = instance['PrivateIpAddress']
                  for tag in instance.get('Tags', []):
                      if tag['Key'] == 'TechTeam':
                         techteam = tag['Value']
                         break
                      else:
                         techteam = 'N/A'
       else:
          instance_name = 'N/A'
       savings.append({
          'InstanceName': recommendation['instanceName'],
          'InstanceId': recommendation['instanceArn'].split('/')[1],
          'CurrentInstanceType': recommendation['currentInstanceType'],
          'FindingReasonCodes': ', '.join(recommendation['findingReasonCodes']),
          'CurrentMaxCPU': recommendation['utilizationMetrics'][0]['value'],
          'TechTeam': techteam,
          'Private Ip Address': Prvipaddr,
          'Instance State': recommendation['instanceState'],
          'Instance Recommendation': ', '.join(item['instanceType'] for item in recommendation['recommendationOptions']),
          'Saving Percentage': ', '.join(str(option['savingsOpportunity']['savingsOpportunityPercentage']) for option in recommendation['recommendationOptions'] if 'savingsOpportunity' in option),
          'Extimated Monthly Savings': ', '.join(str(option['savingsOpportunity']['estimatedMonthlySavings']['value']) for option in recommendation['recommendationOptions'] if 'savingsOpportunity' in option)
       })

over_inst = savings

#Underprovisioned Instances
client = boto3.client('compute-optimizer', region_name=region_name)
ec2_client = boto3.client('ec2', region_name=region_name)

response1 = client.get_ec2_instance_recommendations(
)

under = []

if 'instanceRecommendations' in response1:
   for recommendation in response1['instanceRecommendations']:
       if recommendation['finding'] == 'UNDER_PROVISIONED':
          under.append(recommendation['instanceArn'])

under_savings = []
response = client.get_ec2_instance_recommendations(
    instanceArns=under,
    maxResults=123,
    recommendationPreferences={
        'cpuVendorArchitectures': [
            'AWS_ARM64',
            'CURRENT',
        ]
    }
)

if 'instanceRecommendations' in response:
   for recommendation in response['instanceRecommendations']:
       instance_id = recommendation['instanceArn'].split('/')[1]
       instance_info = ec2_client.describe_instances(InstanceIds=[instance_id])
       if 'Reservations' in instance_info:
          for reservation in instance_info['Reservations']:
              for instance in reservation['Instances']:
                  Prvipaddr = instance.get('PrivateIpAddress', 'Key not found')
                  for tag in instance.get('Tags', []):
                      if tag['Key'] == 'TechTeam':
                         techteam = tag['Value']
                         break
                      else:
                         techteam = 'N/A'
       else:
          instance_name = 'N/A'
       under_savings.append({
          'InstanceName': recommendation['instanceName'],
          'InstanceId': recommendation['instanceArn'].split('/')[1],
          'CurrentInstanceType': recommendation['currentInstanceType'],
          'FindingReasonCodes': ', '.join(recommendation['findingReasonCodes']),
          'CurrentMaxCPU': recommendation['utilizationMetrics'][0]['value'],
          'TechTeam': techteam,
          'Private Ip Address': Prvipaddr,
          'Instance State': recommendation['instanceState'],
          'Instance Recommendation': ', '.join(item['instanceType'] for item in recommendation['recommendationOptions']),
          'Saving Percentage': ', '.join(str(option['savingsOpportunity']['savingsOpportunityPercentage']) for option in recommendation['recommendationOptions'] if 'savingsOpportunity' in option),
          'Extimated Monthly Savings': ', '.join(str(option['savingsOpportunity']['estimatedMonthlySavings']['value']) for option in recommendation['recommendationOptions'] if 'savingsOpportunity' in option)
       })

under_inst = under_savings

#UnOptimizedVolumes
client = boto3.client('compute-optimizer', region_name=region_name)
ec2_client = boto3.client('ec2', region_name=region_name)

response1 = client.get_ebs_volume_recommendations(
)

unopt = []

if 'volumeRecommendations' in response1:
   for recommendation in response1['volumeRecommendations']:
       if recommendation['finding'] == 'NotOptimized':
          unopt.append(recommendation['volumeArn'])


unopt_savings = []
response = client.get_ebs_volume_recommendations(
    volumeArns=unopt,
    maxResults=123,
)


if 'volumeRecommendations' in response:
   for recommendation in response['volumeRecommendations']:
       volume_id = recommendation['volumeArn'].split('/')[1]
       try:
          volume_name = ""
          volume_techteam = ""
          instance_id = ""
          techteam = ""
          Prvipaddr = ""
          response = ec2_client.describe_volumes(VolumeIds=[volume_id])
          if 'Tags' in response['Volumes'][0]:
             for tag in response['Volumes'][0]['Tags']:
                 if tag['Key'] == 'Name':
                     volume_name = tag['Value']
                 if tag['Key'] == 'TechTeam':
                     volume_techteam = tag['Value']

          if 'Attachments' in response['Volumes'][0]:
             instance_id = response['Volumes'][0]['Attachments'][0]['InstanceId']
             instance_info = ec2_client.describe_instances(InstanceIds=[instance_id])
             if 'Reservations' in instance_info:
                for reservation in instance_info['Reservations']:
                    for instance in reservation['Instances']:
                        Prvipaddr = instance['PrivateIpAddress']
                        for tag in instance.get('Tags', []):
                            if tag['Key'] == 'TechTeam':
                               techteam = tag['Value']
                               break
                            else:
                               techteam = 'N/A'
          unopt_savings.append({
             'VolumeName': volume_name,
             'VolumeId': recommendation['volumeArn'].split('/')[1],
             'AttachedToInstanceId': instance_id,
             'VolumeTechTeam': volume_techteam,
             'CurrentVolSize': recommendation['currentConfiguration']['volumeSize'],
             'CurrentVolType': recommendation['currentConfiguration']['volumeType'],
             'CurrentVolIOPS': recommendation['currentConfiguration']['volumeBaselineIOPS'],
             'CurrentVolThroughput': recommendation['currentConfiguration']['volumeBaselineThroughput'],
             'IsCurrentRootVolume': recommendation['currentConfiguration']['rootVolume'],
             'Recommendation Vol IOPS': ', '.join(str(item['configuration']['volumeBaselineIOPS']) for item in recommendation['volumeRecommendationOptions']),
             'Recommendation Vol Size': ', '.join(str(item['configuration']['volumeSize']) for item in recommendation['volumeRecommendationOptions']),
             'Recommendation Vol Throughput': ', '.join(str(item['configuration']['volumeBaselineThroughput']) for item in recommendation['volumeRecommendationOptions']),
             'Attached Instance TechTeam': techteam,
             'Attached Instance Private Ip Address': Prvipaddr,
             'Saving Percentage': ', '.join(str(option['savingsOpportunity']['savingsOpportunityPercentage']) for option in recommendation['volumeRecommendationOptions'] if 'savingsOpportunity' in option),
             'Monthly Savings': ', '.join(str(option['savingsOpportunity']['estimatedMonthlySavings']['value']) for option in recommendation['volumeRecommendationOptions'] if 'savingsOpportunity' in option)
          })
       except ClientError as e:
           if e.response['Error']['Code'] == 'InvalidVolume.NotFound':
               print(f"Volume {volume_id} does not exist.")
           else:
               # Handle other errors if necessary
               print(f"An error occurred: {e}")

unoptimized_volumes = unopt_savings

#Zero IO volumes
ec2_client = boto3.client('ec2', region_name=region_name)
cloudwatch_client = boto3.client('cloudwatch', region_name=region_name)

# Get all EBS volumes in your account
response = ec2_client.describe_volumes()

# Initialize a list to store the volumes with zero IO operations
volumes_with_zero_io = []

# Get the current time
end_time = datetime.now(timezone.utc)
# Calculate the start time (e.g., 7 days ago)
start_time = end_time - timedelta(days=7)  # Adjust the time frame as needed

# Define the metric name for EBS volume read operations (DiskReadOps) and write operations (DiskWriteOps)
metric_names = ['VolumeReadOps', 'VolumeWriteOps']

# Iterate through the volumes
for volume in response['Volumes']:
    volume_id = volume['VolumeId']
    # Iterate through each metric name
    for metric_name in metric_names:
        # Get metric data for the volume
        response = cloudwatch_client.get_metric_statistics(
            Namespace='AWS/EBS',
            MetricName=metric_name,
            Dimensions=[
                {
                    'Name': 'VolumeId',
                    'Value': volume_id
                },
            ],
            StartTime=start_time,
            EndTime=end_time,
            Period=86400,  # 1 hour intervals
            Statistics=['Sum'],
        )
        # Check if there are data points and if the sum is zero for all data points
        if 'Datapoints' in response:
            datapoints = response['Datapoints']
            if all(datapoint['Sum'] == 0 for datapoint in datapoints):
                volumes_with_zero_io.append(volume_id)

unique_list = list(set(volumes_with_zero_io))
                
zero_iops = unique_list


# Initialize the Boto3 EC2 client
ec2_client = boto3.client('ec2', region_name=region_name)  # Replace with your AWS region

# Get a list of all EBS volumes in your account
response = ec2_client.describe_volumes(VolumeIds=zero_iops)

# Initialize a list to store volume information
volume_info_list = []

# Iterate through the volumes
for volume in response['Volumes']:
    volume_id = volume['VolumeId']
    volume_name = ""
    volume_techteam = ""
    instance_id = ""

    # Check if the volume has tags
    if 'Tags' in volume:
        for tag in volume['Tags']:
            if tag['Key'] == 'Name':
                volume_name = tag['Value']
            if tag['Key'] == 'TechTeam':
                volume_techteam = tag['Value']

    # Check if the volume is attached to an instance
    if 'Attachments' in volume and volume['Attachments']:
        instance_id = volume['Attachments'][0]['InstanceId']

        # Get the instance's name by describing the instance
       # Initialize instance_name with a default value
instance_name = 'N/A'

# Get the instance's name by describing the instance
instance_info = ec2_client.describe_instances(InstanceIds=[instance_id])
if 'Reservations' in instance_info:
    for reservation in instance_info['Reservations']:
        for instance in reservation['Instances']:
            for tag in instance.get('Tags', []):
                if tag['Key'] == 'Name':
                    instance_name = tag['Value']
                    break

    # Append the volume information to the list
    volume_info_list.append({
        'Volume ID': volume_id,
        'Volume Name': volume_name,
        'Instance ID': instance_id,
        'Instance Name': instance_name,
        'Volume Type': volume['VolumeType'],
        'Size in GiBs': volume['Size'],
        'Provisioned IOPS': volume['Iops'],
        'TechTeam': volume_techteam
    })

zero_io_vol = volume_info_list

#ASG Report
client = boto3.client('autoscaling', region_name=region_name)
client1 = boto3.client('cloudwatch', region_name=region_name)
response = client.describe_auto_scaling_groups()
end_time = datetime.now(timezone.utc)
start_time = end_time - timedelta(days=12)
ASGs=[]
for ASG in response['AutoScalingGroups']:
     techtag = next((tag['Value'] for tag in ASG.get('Tags', []) if tag['Key'] == 'TechTeam'), None)
     scheduled_actions = client.describe_scheduled_actions(AutoScalingGroupName=ASG['AutoScalingGroupName'])
     result = "NO" if len(scheduled_actions['ScheduledUpdateGroupActions']) == 0 else "YES"
     asg_name = ASG['AutoScalingGroupName']
     response = client1.get_metric_statistics(
        Namespace='AWS/EC2',
        MetricName='CPUUtilization',
        Dimensions=[
            {
                'Name': 'AutoScalingGroupName',
                'Value': asg_name,
            },
        ],
        StartTime=start_time,
        EndTime=end_time,
        Period=3600,  # 1-hour intervals
        Statistics=['Maximum', 'Average']
     )
     if 'Datapoints' in response and len(response['Datapoints']) > 0:
        max_cpu = max(response['Datapoints'], key=lambda x: x['Maximum'])['Maximum']
        avg_cpu = sum([datapoint['Average'] for datapoint in response['Datapoints']]) / len(response['Datapoints'])
     else:
        max_cpu = None
        avg_cpu = None
     ASGs.append({
         'ASGNAME': ASG['AutoScalingGroupName'],
         'DesiredSize': ASG['DesiredCapacity'],
         'MinSize': ASG['MinSize'],
         'MaxSize': ASG['MaxSize'],
         'CreatedTime': ASG['CreatedTime'].astimezone(None).strftime('%Y-%m-%d %H:%M:%S'),
         'TechTeam': techtag,
         'Scheduled Action' : result,
         'MaximumCPUUtilization': max_cpu,
         'AverageCPUUtilization': avg_cpu
         
     })
asg_report = ASGs

#S3 with NO Lifecycle POlicy
#s3_client = boto3.client('s3')

# List buckets
#buckets = s3_client.list_buckets()['Buckets']

# Initialize a list to store bucket names with no lifecycle policy
#buckets_with_no_lifecycle = []

# Check each bucket for a lifecycle policy
#for bucket in buckets:
#    bucket_name = bucket['Name']
#    
#    try:
#        # Attempt to get the bucket's lifecycle configuration
#        response = s3_client.get_bucket_lifecycle(Bucket=bucket_name)
#    except ClientError as e:
#        if e.response['Error']['Code'] == 'NoSuchLifecycleConfiguration':
#            # If NoSuchLifecycleConfiguration exception is raised, the bucket has no lifecycle policy
#            try:
#                bucket_name = bucket['Name']
#
#            # Get tags for the bucket
#                tags = s3_client.get_bucket_tagging(Bucket=bucket_name).get('TagSet', [])
#
#            # Check if the 'Techteam' tag is present
#                tech_tag_value = None
#                for tag in tags:
#                    if tag['Key'] == 'TechTeam':
#                        tech_tag_value = tag['Value']
#                        break
#                 
#                buckets_with_no_lifecycle.append({
#                         'BucketName': bucket_name,
#                         'TechTeam': tech_tag_value
#                })
#
#            # Write bucket name and tag value (or 'NONE') to the Excel sheet
#            except Exception as e:
#                # Handle Access Denied errors gracefully
#                if "AccessDenied" in str(e):
#                    print(f"Access Denied for Bucket: {bucket['Name']}")
#                    continue
#                else:
#                    # Handle other exceptions as needed
#                    print(f"Error for Bucket: {bucket_info['Name']}\n{str(e)}")
#        elif e.response['Error']['Code'] == 'NoSuchBucket':
#            # Handle NoSuchBucket exception if the bucket no longer exists
#            continue
#        nolifecycle = buckets_with_no_lifecycle
#
#s3_no_lifecycle_policy = nolifecycle

#CloudWatch LOg Groups with newver expire retention
cloudwatch_logs = boto3.client('logs', region_name=region_name)

# List all log groups
response = cloudwatch_logs.describe_log_groups()

log_groups = []

# Iterate through the log groups and check the retention settings
for log_group in response['logGroups']:
    retention_in_days =log_group.get('retentionInDays', -1)
    if retention_in_days == -1:
       log_groups.append({
          'Log_Group_Name': log_group['logGroupName'],
          'Retention': 'Never Expire'
    })

never_expire_log_groups = log_groups

#NonAMDNonGraviton
ec2_client = boto3.client('ec2', region_name=region_name)  # Replace with your AWS region

# Get a list of all EC2 instances in your account
response = ec2_client.describe_instances()

# Initialize a list to store instances with numbers followed by 'a' or 'g'
instances_with_nonamd_nongraviton = []

# Regular expression pattern to match numbers followed by 'a' or 'g'
pattern = r'^(?:(?!\d+[agAG]\.).)*$'

# Iterate through the instances
for reservation in response['Reservations']:
    for instance in reservation['Instances']:
        instance_id = instance['InstanceId']
        instance_type = instance['InstanceType']

        # Check if the instance type contains a number followed by 'a' or 'g'
        if re.search(pattern, instance_type):
            instances_with_nonamd_nongraviton.append({
                'Instance ID': instance_id,
                'Instance Type': instance_type,
                'Ip': instance.get('PrivateIpAddress', 'Key not found'),
                'Instance Type': instance['InstanceType'],
                'Instance State': instance['State']['Name'],
                'Availability Zone': instance['Placement']['AvailabilityZone'],
                'Name': next((tag['Value'] for tag in instance.get('Tags', []) if tag['Key'] == 'Name'), None),
                'Techteam': next((tag['Value'] for tag in instance.get('Tags', []) if tag['Key'] == 'Techteam'), None),
                'Environment': next((tag['Value'] for tag in instance.get('Tags', []) if tag['Key'] == 'Environment'), None)
                
            })

nonamdgraviton = instances_with_nonamd_nongraviton

#EIP
client = boto3.client('ec2', region_name=region_name)
response = client.describe_addresses()
eips=[]
for address in response['Addresses']:
    if 'InstanceId' not in address:
        eips.append({
           'PublicIP': address['PublicIp']
        })
unused_eips = eips


#Volumes
client = boto3.client('ec2', region_name=region_name)
response=client.describe_volumes()

volumes = []
for volume in response['Volumes']:
    if len(volume['Attachments']) == 0:
        techtag = next((tag['Value'] for tag in volume.get('Tags', []) if tag['Key'] == 'Techteam'), None)
        volumes.append({
            'VolumeId':  volume['VolumeId'],
            'VolumeType': volume['VolumeType'],
            'VolumeSize (GiB)':  volume['Size'],
            'IOPS': volume['Iops'],
            'Throughput': volume.get('Throughput', 'Key not found'),
            'Snapshot': volume['SnapshotId'],
            'CreateTime': volume['CreateTime'].astimezone(None).strftime('%Y-%m-%d %H:%M:%S'),
            'TechTeam': techtag
        })
unused_volumes = volumes
      


#Snapshots
# Make a list of existing volumes
volume_response = client.describe_volumes()
volumes = [volume['VolumeId'] for volume in volume_response['Volumes']]

# Find snapshots without existing volume
snapshots = []
snapshot_response = client.describe_snapshots(OwnerIds=['self'])
for snapshot in snapshot_response['Snapshots']:
    if snapshot['VolumeId'] not in volumes:
        techtag = next((tag['Value'] for tag in snapshot.get('Tags', []) if tag['Key'] == 'Techteam'), None)
        snapshots.append({
            'SnapshotId': snapshot['SnapshotId'],
            'SnapshotStartTime': snapshot['StartTime'].astimezone(None).strftime('%Y-%m-%d %H:%M:%S'),
            'SnapshotSize': snapshot['VolumeSize'],
            'Description': snapshot['Description'],
            'TechTeam': techtag
        })
unused_snapshots = snapshots

#Ami's
client = boto3.client('ec2', region_name=region_name)

instances = client.describe_instances()
used_amis = []
for reservation in instances['Reservations']:
    for instance in reservation['Instances']:
        used_amis.append(instance['ImageId'])

custom_images = client.describe_images(
    Filters=[
            {
            'Name': 'state',
            'Values': [
                'available'
            ]
        },
    ],
    Owners= ['self']
)

custom_amis_list = []

for image in custom_images['Images']:
    if image['ImageId'] not in used_amis:
       techtag = next((tag['Value'] for tag in image.get('Tags', []) if tag['Key'] == 'Techteam'), None)
       custom_amis_list.append({
         'AMIID': image['ImageId'],
         'Name': image['Name'],
         'Creationdate': image['CreationDate'],
         'TechTeam': techtag
       })
unused_ami = custom_amis_list

#Securit Groups

sginuse = client.describe_network_interfaces()
sec_groups_in_use = []
for SG in sginuse['NetworkInterfaces']:
    for SGinuse in SG['Groups']:
        sec_groups_in_use.append(SGinuse['GroupId'])


#unused_sec_groups = []

response = client.describe_security_groups()
all_sec_groups = []
for SecGrp in response['SecurityGroups']:
    if SecGrp['GroupId'] not in sec_groups_in_use:
       techtag = next((tag['Value'] for tag in SecGrp.get('Tags', []) if tag['Key'] == 'Techteam'), None)
       all_sec_groups.append({
          'SGID': SecGrp['GroupId'],
          'GroupName': SecGrp['GroupName'],
          'Description': SecGrp['Description'],
          'TechTeam': techtag
       })

unused_sgs = all_sec_groups

#VPC
client = boto3.client('ec2', region_name=region_name)
vpc_response = client.describe_vpcs()
#vpcs = [vpc['VpcId'] for vpc in vpc_response['Vpcs']]
vpcs = [vpc for vpc in vpc_response['Vpcs']]
unusedvpc = []
for VPC in vpcs:
    eni = client.describe_network_interfaces(
            Filters=[
                {
                    'Name': 'vpc-id',
                    'Values': [VPC['VpcId']]
                    }
            ])
    enis = [NI['NetworkInterfaceId'] for NI in eni['NetworkInterfaces']]
    if len(enis) == 0:
        techtag = next((tag['Value'] for tag in VPC.get('Tags', []) if tag['Key'] == 'Techteam'), None)
        unusedvpc.append({
            'VPCID': VPC['VpcId'],
            'CidrBlock': VPC['CidrBlock'],
            'TechTeam': techtag
        })
unused_vpcs = unusedvpc

#ELBv2
client = boto3.client('elbv2', region_name=region_name)
        
response_targets = client.describe_target_groups()
response_lbs = client.describe_load_balancers()
        
elbsv2_arns=[]
elbsv2_target_arns=[]
elbsv2_notinuse=[]    
elbsv2_unused=[]
elbsv2_unhealthy=[]
elbsv2_nolisteners=[]


#ELBv2 check Target groups
for tg in response_targets['TargetGroups']:
        elbsv2_target_arns.append(tg['TargetGroupArn'])
  

for target_group in elbsv2_target_arns:
    response_healthy = client.describe_target_health(TargetGroupArn=target_group)
    for target in response_healthy['TargetHealthDescriptions']:
        if target['TargetHealth']['State'] == "unused" :
            elbsv2_unused.append(target_group)
        elif target['TargetHealth']['State'] == "unhealthy" :
            elbsv2_unhealthy.append(target_group)

for tgunused in response_targets['TargetGroups']:
    response_tags = client.describe_tags(ResourceArns=[tgunused['TargetGroupArn']])
    tghealth = client.describe_target_health(TargetGroupArn=tgunused['TargetGroupArn'])
    unused = [TG for TG in tghealth['TargetHealthDescriptions']]
    if len(unused) == 0:
        techtag = next((tag['Value'] for tag in response_tags['TagDescriptions'][0]['Tags'] if tag['Key'] == 'Techteam'), None)
        elbsv2_notinuse.append({
            'TGNAME': tgunused['TargetGroupArn'].split('/')[1],
            'LBName': "\n".join(tgunused['LoadBalancerArns']),
            'TechTeam': techtag
        })

unused_tgs = elbsv2_notinuse
unusedlb = []
#ELBv2 check Load Balancers without Listeners
for ELB in response_lbs['LoadBalancers']:
    elbsv2_arns.append(ELB['LoadBalancerArn'])

for elbtg in response_lbs['LoadBalancers']:
    response_tags = client.describe_tags(ResourceArns=[elbtg['LoadBalancerArn']])
    tgname = client.describe_target_groups(LoadBalancerArn=elbtg['LoadBalancerArn'])
    elbtgname = [TG['TargetGroupName'] for TG in tgname['TargetGroups']]
    check =  all(item in elbsv2_notinuse for item in elbtgname)
    if check is True:
        techtag = next((tag['Value'] for tag in response_tags['TagDescriptions'][0]['Tags'] if tag['Key'] == 'Techteam'), None)
        unusedlb.append({
            'LBNAME': elbtg['LoadBalancerArn'].split('/')[2],
            'Scheme': elbtg['Scheme'],
            'Type': elbtg['Type'],
            'CreatedTime': elbtg['CreatedTime'].astimezone(None).strftime('%Y-%m-%d %H:%M:%S'),
            'TechTeam': techtag
        })
unused_elbsv2 = unusedlb

unusedstatetg = []

        
for ELB in elbsv2_arns:
    response_listeners = client.describe_listeners(LoadBalancerArn=ELB) 
    if response_listeners['Listeners'] == []:
        elbsv2_nolisteners.append(ELB)
        
# print ELBsv2
if len(elbsv2_unused) > 0:
    for elbv2 in elbsv2_unused:
        unusedstatetg.append({
            'UnusedStateTGName': elbv2.split('/')[1]
        })

unused_state_tgs = [dict(t) for t in {tuple(d.items()) for d in unusedstatetg}]

unhealthystatetg = []
if len(elbsv2_unhealthy) > 0:
    for elbv2 in elbsv2_unhealthy:
        unhealthystatetg.append({
             'UnhealthyStateTGName': elbv2.split('/')[1]
        })

unhealthy_state_tgs = [dict(t) for t in {tuple(d.items()) for d in unhealthystatetg}]

nolistenerslbs = []
        
if len(elbsv2_nolisteners) > 0:
    for elbv2 in elbsv2_nolisteners:
        nolistenerslbs.append({
             'NOListenerALBName': elbv2
        })

nolisteners_albs = nolistenerslbs

#ELBv1
client = boto3.client('elb', region_name=region_name)
response = client.describe_load_balancers()
elbs=[]
for ELB in response['LoadBalancerDescriptions']:
    if len(ELB['Instances']) == 0:
        elbs.append({
           'LBNAME': ELB['LoadBalancerName']
        })
unused_clbs = elbs

#Autoscaling
client = boto3.client('autoscaling', region_name=region_name)
response = client.describe_launch_configurations()
LC_list=[]
for LC in response['LaunchConfigurations']:
    LC_name = LC['LaunchConfigurationName']
    LC_list.append(LC_name)
response1 = client.describe_auto_scaling_groups()
for ASG in response1['AutoScalingGroups']:
    if ASG.get('LaunchConfigurationName') in LC_list:
                    LC_list.remove(ASG['LaunchConfigurationName'])
LCs=[]
for LC in LC_list:
    LCs.append({
        'LaunchConfigurationName': LC
    })
unused_lc = LCs


response = client.describe_auto_scaling_groups()
ASGs=[]
for ASG in response['AutoScalingGroups']:
    if ASG['DesiredCapacity'] == 0:
        techtag = next((tag['Value'] for tag in ASG.get('Tags', []) if tag['Key'] == 'Techteam'), None)
        ASGs.append({
            'ASGNAME': ASG['AutoScalingGroupName'],
            'MinSize': ASG['MinSize'],
            'MaxSize': ASG['MaxSize'],
            'CreatedTime': ASG['CreatedTime'].astimezone(None).strftime('%Y-%m-%d %H:%M:%S'),
            'TechTeam': techtag
        })
unused_asgs = ASGs


ec2_client = boto3.client('ec2', region_name=region_name)
launch_templates = ec2_client.describe_launch_templates()['LaunchTemplates']
lt_ids = [template['LaunchTemplateId'] for template in launch_templates]
autoscaling_client = boto3.client('autoscaling', region_name=region_name)
asg_descriptions = autoscaling_client.describe_auto_scaling_groups()['AutoScalingGroups']
used_launch_template_ids = []
used_lt = []
for asg_description in asg_descriptions:
    launch_template = asg_description.get('LaunchTemplate', {})
    launch_template_id = launch_template.get('LaunchTemplateId')
    if launch_template_id is None:
       launch_template1 = asg_description.get('MixedInstancesPolicy', {})
       launch_template_id1 = launch_template1.get('LaunchTemplate', {})
       launch_template_id2 = launch_template_id1.get('LaunchTemplateSpecification', {})
       launch_template_id3 = launch_template_id2.get('LaunchTemplateId')
       used_lt.append(launch_template_id3)
    else:
       used_launch_template_ids.append(launch_template_id)
used_lts = list(set(used_lt + used_launch_template_ids))
unused_lt = [item for item in lt_ids if item not in used_lts]
unused_lts = []
for i in unused_lt:
    response = ec2_client.describe_launch_templates(LaunchTemplateIds=[i])
    unused_lts.append({
          'LaunchTemplateID': i,
          'LaunchTemplateName': response['LaunchTemplates'][0]['LaunchTemplateName'],
          'CreateTime': response['LaunchTemplates'][0]['CreateTime'].astimezone(None).strftime('%Y-%m-%d %H:%M:%S')
    })
unused_launch_templates_ids = unused_lts
    



if len(unused_volumes) != 0:
   volumes_sheet = wb.create_sheet(title='Unused Volumes')
if len(unused_eips) != 0:
   eips_sheet = wb.create_sheet(title='Unused EIPS')
if len(unused_snapshots) != 0:
   snapshots_sheet = wb.create_sheet(title='Unused Snapshots')
if len(unused_ami) != 0:
   amis_sheet = wb.create_sheet(title='Unused AMIS')
if len(unused_sgs) != 0:
   sg_sheet = wb.create_sheet(title='Unused Security Groups')
if len(unused_vpcs) != 0:
   vpc_sheet = wb.create_sheet(title='Unused VPCS')
if len(unused_tgs) != 0:
   tgs_sheet = wb.create_sheet(title='Unused TGS')
if len(unused_elbsv2) != 0:
   alb_sheet = wb.create_sheet(title='Unused ELBV2S')
if len(unused_state_tgs) != 0:
   unused_state_tgs_sheet = wb.create_sheet(title='Unused State TGS')
if len(unhealthy_state_tgs) != 0:
   unhealthy_state_tgs_sheet = wb.create_sheet(title='Unhealthy State TGS')
if len(nolisteners_albs) != 0:
   nolisteners_alb_sheet = wb.create_sheet(title='No Listeners ALBS')
if len(unused_clbs) != 0:
   clb_sheet = wb.create_sheet(title='Unused ELBV1S')
if len(unused_lc) != 0:
   lc_sheet = wb.create_sheet(title='Unused LCS')
if len(unused_asgs) != 0:
   asg_sheet = wb.create_sheet(title='Unused ASGS')
if len(unused_launch_templates_ids) != 0:
   lt_sheet = wb.create_sheet(title='Unused LTS')
if len(stopint) != 0:
   stopped_instances_sheet = wb.create_sheet(title='Stopped Instances')
if len(over_inst) != 0:
   overprovisioned_sheet = wb.create_sheet(title='OverProvisioned Instances')
if len(under_inst) != 0:
   underprovisioned_sheet = wb.create_sheet(title='UnderProvisioned Instances')
if len(unoptimized_volumes) != 0:
   unoptimizedvolumes_sheet = wb.create_sheet(title='UnOptimized Volumes')
if len(zero_io_vol) != 0:
   zeroio_sheet = wb.create_sheet(title='Zero Io Volumes')
if len(asg_report) != 0:
   asgreport_sheet = wb.create_sheet(title='ASG Report')
#if len(s3_no_lifecycle_policy) != 0:
#   s3nolifecycle_sheet = wb.create_sheet(title='S3 Bucket With No Lifecycle Policy')
if len(never_expire_log_groups) != 0:
   cloudwatch_loggroup_sheet = wb.create_sheet(title='CloudWatch LogGroup With Never Expire Retention')
if len(nonamdgraviton) != 0:
   non_amd_non_graviton_sheet = wb.create_sheet(title='Non AMD and Graviton Instances')
#if len(npnns) != 0:
#   non_prod_infra_shutdown_sheet = wb.create_sheet(title='Non prod Infra Shutdown')

bold_font = Font(bold=True)

def populate_sheet(sheet, data):
    sheet.append(list(data[0].keys()))
    for cell in sheet[1]:
        cell.font = bold_font
    for item in data:
        sheet.append(list(item.values()))

if len(unused_volumes) != 0:
   populate_sheet(volumes_sheet, unused_volumes)
if len(unused_eips) != 0:
   populate_sheet(eips_sheet, unused_eips)
if len(unused_snapshots) != 0:
   populate_sheet(snapshots_sheet, unused_snapshots)
if len(unused_ami) != 0:
   populate_sheet(amis_sheet, unused_ami)
if len(unused_sgs) != 0:
   populate_sheet(sg_sheet, unused_sgs)
if len(unused_vpcs) != 0:
   populate_sheet(vpc_sheet, unused_vpcs)
if len(unused_tgs) != 0:
   populate_sheet(tgs_sheet, unused_tgs)
if len(unused_elbsv2) != 0:
   populate_sheet(alb_sheet, unused_elbsv2)
if len(unused_state_tgs) != 0:
   populate_sheet(unused_state_tgs_sheet, unused_state_tgs)
if len(unhealthy_state_tgs) != 0:
   populate_sheet(unhealthy_state_tgs_sheet, unhealthy_state_tgs)
if len(nolisteners_albs) != 0:
   populate_sheet(nolisteners_alb_sheet, nolisteners_albs)
if len(unused_clbs) != 0:
   populate_sheet(clb_sheet, unused_clbs)
if len(unused_lc) != 0:
   populate_sheet(lc_sheet, unused_lc)
if len(unused_asgs) != 0:
   populate_sheet(asg_sheet, unused_asgs)
if len(unused_launch_templates_ids) != 0:
   populate_sheet(lt_sheet, unused_launch_templates_ids)
if len(stopint) != 0:
   populate_sheet(stopped_instances_sheet, stopint)
if len(over_inst) != 0:
   populate_sheet(overprovisioned_sheet, over_inst)
if len(under_inst) != 0:
   populate_sheet(underprovisioned_sheet, under_inst)
if len(unoptimized_volumes) != 0:
   populate_sheet(unoptimizedvolumes_sheet, unoptimized_volumes)
if len(zero_io_vol) != 0:
   populate_sheet(zeroio_sheet, zero_io_vol)
if len(asg_report) != 0:
   populate_sheet(asgreport_sheet, asg_report)
#if len(s3_no_lifecycle_policy) != 0:
#   populate_sheet(s3nolifecycle_sheet, s3_no_lifecycle_policy)
if len(never_expire_log_groups) != 0:
   populate_sheet(cloudwatch_loggroup_sheet, never_expire_log_groups)
if len(nonamdgraviton) != 0:
   populate_sheet(non_amd_non_graviton_sheet, nonamdgraviton)
#if len(npnns) != 0:
#   populate_sheet(non_prod_infra_shutdown_sheet, npnns)

del wb['Sheet']

wb.save(file_name)
